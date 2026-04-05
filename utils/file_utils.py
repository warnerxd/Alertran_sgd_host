# utils/file_utils.py
"""
Utilidades para manejo de archivos
"""
from pathlib import Path
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from typing import List, Optional

class FileUtils:
    """Clase utilitaria para operaciones con archivos"""
    
    @staticmethod
    def obtener_carpeta_descargas() -> Path:
        """Obtiene la ruta de la carpeta de Descargas del usuario"""
        home = Path.home()
        
        if os.name == 'nt':  # Windows
            downloads = home / 'Downloads'
            if downloads.exists():
                return downloads
        
        # macOS/Linux
        downloads = home / 'Descargas'
        if downloads.exists():
            return downloads
        
        downloads = home / 'Downloads'
        if downloads.exists():
            return downloads
        
        descargas_dir = home / 'Descargas_Alertran'
        descargas_dir.mkdir(exist_ok=True)
        return descargas_dir

    @staticmethod
    def generar_nombre_unico(carpeta: Path, base_nombre: str, extension: str) -> Path:
        """Genera un nombre de archivo único"""
        contador = 1
        nombre_archivo = carpeta / f"{base_nombre}.{extension}"
        
        while nombre_archivo.exists():
            nombre_archivo = carpeta / f"{base_nombre}_{contador}.{extension}"
            contador += 1
        
        return nombre_archivo

    @staticmethod
    def leer_guias_excel(ruta: Path) -> List[str]:
        """Lee el archivo Excel y extrae las guías"""
        wb = load_workbook(ruta, read_only=True, data_only=True)
        try:
            ws = wb.active
            guias = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    guia = str(row[0]).strip()
                    if guia:
                        guias.append(guia)
            return guias
        finally:
            wb.close()

    @staticmethod
    def guardar_errores_excel(guias_error: List, guias_advertencia: List, carpeta: Path) -> Optional[str]:
        """Guarda el archivo de errores"""
        if not guias_error and not guias_advertencia:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores y Advertencias"
        
        ws.append(["Guía", "Motivo", "Tipo", "Fecha/Hora"])
        
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for guia, motivo in guias_error:
            ws.append([guia, motivo, "ERROR", fecha_actual])
        
        for guia, motivo in guias_advertencia:
            ws.append([guia, motivo, "ADVERTENCIA", fecha_actual])
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_nombre = f"errores_alertran_{timestamp}"
        ruta_archivo = FileUtils.generar_nombre_unico(carpeta, base_nombre, "xlsx")
        
        wb.save(ruta_archivo)
        return str(ruta_archivo)

    @staticmethod
    def guardar_log(log_contenido: str, carpeta: Path) -> str:
        """Guarda el log completo"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_nombre = f"log_alertran_{timestamp}"
        ruta_archivo = FileUtils.generar_nombre_unico(carpeta, base_nombre, "txt")
        
        with open(ruta_archivo, 'w', encoding='utf-8') as f:
            f.write(log_contenido)
        
        return str(ruta_archivo)