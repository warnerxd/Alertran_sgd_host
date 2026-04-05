# ui/historial_window.py
"""
Ventana de historial de guías procesadas con estadísticas y filtros múltiples
Actualización en tiempo real de estadísticas circulares
"""

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QMessageBox, QApplication, QWidget, QCheckBox, QGroupBox,
    QGridLayout, QFrame, QSizePolicy, QSplitter, QMenu
)
from ui.widgets.progress_bar import MacProgressBar
from PySide6.QtGui import QFont, QColor, QPainter, QBrush, QPen
from PySide6.QtCore import Qt, QTimer, Signal, QRect, QByteArray
from utils.settings_manager import SettingsManager
from datetime import datetime, timedelta
from pathlib import Path

from utils.file_utils import FileUtils
from utils import theme

class EstadisticasWidget(QWidget):
    """Widget circular para mostrar estadísticas con actualización en tiempo real"""
    
    def __init__(self, titulo, valor, color, parent=None):
        super().__init__(parent)
        self.titulo = titulo
        self.valor = valor
        self.color = color
        self.setMinimumSize(70, 80)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        w, h = self.width(), self.height()
        titulo_h = 18
        margin = 6
        circle_size = min(w - margin * 2, h - titulo_h - margin * 2)
        x = (w - circle_size) // 2
        y = margin

        rect = QRect(x, y, circle_size, circle_size)

        pen = QPen(QColor(self.color), 2)
        painter.setPen(pen)

        c = theme.colors()
        if self.valor > 0:
            light_factor = 160 if theme.is_dark() else 140
            brush = QBrush(QColor(self.color).lighter(light_factor))
            painter.setBrush(brush)
            painter.drawEllipse(rect)

            font_size = max(8, circle_size // 4)
            font = QFont("Arial", font_size, QFont.Weight.Bold)
            painter.setFont(font)
            painter.setPen(QColor(self.color))
            painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, str(self.valor))
        else:
            brush = QBrush(QColor(c['stat_zero']))
            painter.setBrush(brush)
            painter.drawEllipse(rect)

        font = QFont("Arial", 7)
        painter.setFont(font)
        painter.setPen(QColor(c['stat_text']))
        titulo_rect = QRect(0, y + circle_size + 2, w, titulo_h)
        painter.drawText(titulo_rect, Qt.AlignmentFlag.AlignCenter, self.titulo)
    
    def actualizar_valor(self, nuevo_valor):
        """Actualiza el valor del círculo y fuerza el redibujado"""
        if self.valor != nuevo_valor:
            self.valor = nuevo_valor
            self.update()  # Forzar repintado inmediato

class HistorialWindow(QDialog):
    """Ventana para mostrar el historial de guías procesadas"""
    
    # Señal para cuando se actualizan los filtros
    filtros_actualizados = Signal()
    # Señal para solicitar actualización al padre
    solicitar_actualizacion_signal = Signal()
    # Señal para reprocesar una guía individual desde el historial
    reprocesar_guia = Signal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📊 ANÁLISIS DE GUÍAS PROCESADAS")
        self.setMinimumSize(600, 420)
        self.resize(1050, 640)
        self.setSizeGripEnabled(True)
        self.setModal(True)
        _geom = SettingsManager.get_instance().get("historial_window_geometry", None)
        if _geom:
            self.restoreGeometry(QByteArray.fromBase64(_geom.encode()))
        
        self.datos_completos = []
        self.datos_filtrados = []
        self.total_esperado = 0   # total de guías a procesar (pasado desde main)
        self.filtros_activos = {
            "✅": False,
            "📦": False, 
            "❌": False,
            "⚠️": False
        }
        self.texto_busqueda = ""
        self.navegador_filtro = "Todos"
        self.actualizacion_automatica = False
        
        # Variables para el contador de tiempo
        self.tiempo_inicio = datetime.now()
        self.timer_contador = QTimer()
        self.timer_contador.setInterval(1000)  # 1 segundo
        self.timer_contador.timeout.connect(self.actualizar_contador_tiempo)
        self.timer_contador.start()
        
        self.carpeta_descargas = FileUtils.obtener_carpeta_descargas()
        self.file_utils = FileUtils()
        
        # Timer para actualización automática (cada 2 segundos para tiempo real)
        self.timer_actualizacion = QTimer()
        self.timer_actualizacion.setInterval(2000)  # 2 segundos para tiempo real
        self.timer_actualizacion.timeout.connect(self._actualizacion_automatica_timeout)
        
        self._setup_ui()
        self._setup_styles()
        self._conectar_senales()
        from utils import theme
        theme.signals.changed.connect(self._setup_styles)

    def _setup_ui(self):
        """Configura la interfaz de usuario"""
        layout_principal = QHBoxLayout(self)
        layout_principal.setSpacing(0)
        layout_principal.setContentsMargins(10, 10, 10, 10)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.setHandleWidth(6)

        panel_izquierdo = self._crear_panel_izquierdo()
        splitter.addWidget(panel_izquierdo)

        panel_derecho = self._crear_panel_derecho()
        splitter.addWidget(panel_derecho)

        # Proporción inicial: ~22% izquierdo, ~78% derecho — el derecho absorbe todo el resize
        splitter.setSizes([230, 800])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        layout_principal.addWidget(splitter)

    def _crear_panel_izquierdo(self):
        """Crea el panel izquierdo con estadísticas y filtros"""
        panel = QWidget()
        panel.setObjectName("panel_izquierdo")
        panel.setMinimumWidth(180)
        panel.setMaximumWidth(300)

        layout = QVBoxLayout(panel)
        layout.setSpacing(10)

        # Título del panel
        titulo = QLabel("📊 ESTADÍSTICAS")
        titulo.setObjectName("titulo_panel")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setWordWrap(True)
        layout.addWidget(titulo)
        
        # Línea separadora
        linea = QFrame()
        linea.setFrameShape(QFrame.Shape.HLine)
        linea.setObjectName("separador")
        layout.addWidget(linea)
        
        # Widget de estadísticas circulares
        self.stats_widget = QWidget()
        stats_layout = QGridLayout(self.stats_widget)
        stats_layout.setSpacing(2)
        stats_layout.setContentsMargins(2, 2, 2, 2)
        
        # Crear estadísticas circulares con colores vibrantes
        self.stats = {
            "total": EstadisticasWidget("TOTAL", 0, "#3498db"),      # Azul
            "exitosas": EstadisticasWidget("✅ EXITOSAS", 0, "#27ae60"),  # Verde
            "ent": EstadisticasWidget("📦 ENT", 0, "#f39c12"),        # Naranja
            "errores": EstadisticasWidget("❌ ERRORES", 0, "#e74c3c"), # Rojo
            "advertencias": EstadisticasWidget("⚠️ ADVERTENCIAS", 0, "#e67e22") # Naranja oscuro
        }
        
        # Posicionar en grid con alineación centrada
        stats_layout.addWidget(self.stats["total"], 0, 0, Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.stats["exitosas"], 0, 1, Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.stats["ent"], 0, 2, Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.stats["errores"], 1, 0, Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.stats["advertencias"], 1, 1, Qt.AlignmentFlag.AlignCenter)
        
        # Espacio vacío en la posición (1,2)
        spacer = QLabel("")
        stats_layout.addWidget(spacer, 1, 2)
        
        layout.addWidget(self.stats_widget)

        # Panel de filtros
        filtros_group = QGroupBox("🔍 FILTROS POR ESTADO")
        filtros_group.setObjectName("filtros_group")
        filtros_layout = QVBoxLayout(filtros_group)
        
        # Checkbox "Seleccionar todos"
        self.cb_seleccionar_todos = QCheckBox("✅ SELECCIONAR TODOS")
        self.cb_seleccionar_todos.setObjectName("cb_seleccionar_todos")
        self.cb_seleccionar_todos.setStyleSheet("""
            QCheckBox#cb_seleccionar_todos {
                color: #3498db;
                font-weight: bold;
                font-size: 10pt;
                spacing: 8px;
                padding: 5px;
                background-color: #2c3e50;
                border-radius: 3px;
            }
        """)
        filtros_layout.addWidget(self.cb_seleccionar_todos)
        
        # Línea separadora
        linea_filtros = QFrame()
        linea_filtros.setFrameShape(QFrame.Shape.HLine)
        linea_filtros.setStyleSheet(f"background-color: {theme.colors()['sep']}; max-height: 1px;")
        filtros_layout.addWidget(linea_filtros)
        
        # Checkboxes de filtros individuales
        self.checkboxes = {}
        for estado, color, simbolo in [
            ("✅ Exitosas", "#27ae60", "✅"),
            ("📦 ENT", "#f39c12", "📦"),
            ("❌ Errores", "#e74c3c", "❌"),
            ("⚠️ Advertencias", "#e67e22", "⚠️")
        ]:
            cb = QCheckBox(estado)
            cb.setProperty("simbolo", simbolo)  # Guardar el símbolo como propiedad
            cb.setStyleSheet(f"""
                QCheckBox {{
                    color: {color};
                    font-weight: bold;
                    spacing: 8px;
                    padding: 3px;
                }}
                QCheckBox::indicator {{
                    width: 18px;
                    height: 18px;
                }}
                QCheckBox::indicator:checked {{
                    background-color: {color};
                    border: 2px solid white;
                    border-radius: 3px;
                }}
            """)
            self.checkboxes[simbolo] = cb  # Usar símbolo como clave
            filtros_layout.addWidget(cb)
        
        # Filtro por navegador
        filtros_layout.addSpacing(10)
        filtros_layout.addWidget(QLabel("🌐 Navegador:"))
        self.navegador_combo = QComboBox()
        self.navegador_combo.addItems(["Todos"])
        self.navegador_combo.setObjectName("navegador_combo")
        filtros_layout.addWidget(self.navegador_combo)
        
        # Botones de acción — responsive, mismo tamaño
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(6)
        self.btn_aplicar = QPushButton("✅ APLICAR")
        self.btn_aplicar.setObjectName("btn_aplicar_filtros")
        self.btn_aplicar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        btn_layout.addWidget(self.btn_aplicar)

        self.btn_limpiar = QPushButton("🗑 LIMPIAR")
        self.btn_limpiar.setObjectName("btn_limpiar_filtros")
        self.btn_limpiar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        btn_layout.addWidget(self.btn_limpiar)

        filtros_layout.addLayout(btn_layout)
        
        layout.addWidget(filtros_group)
        layout.addStretch()
        
        return panel

    def _crear_panel_derecho(self):
        """Crea el panel derecho con la tabla y controles"""
        panel = QWidget()
        panel.setObjectName("panel_derecho")
        
        layout = QVBoxLayout(panel)
        layout.setSpacing(10)
        
        # Toolbar superior
        toolbar = self._crear_toolbar()
        layout.addLayout(toolbar)
        
        # Barra de búsqueda
        search_layout = QHBoxLayout()
        search_layout.setSpacing(6)

        lbl_buscar = QLabel("🔎 Buscar guía:")
        lbl_buscar.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        search_layout.addWidget(lbl_buscar)

        self.buscar_input = QComboBox()
        self.buscar_input.setEditable(True)
        self.buscar_input.setPlaceholderText("Escriba para buscar...")
        self.buscar_input.setObjectName("buscar_input")
        self.buscar_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        search_layout.addWidget(self.buscar_input, stretch=1)

        self.btn_buscar = QPushButton("🔍 Buscar")
        self.btn_buscar.setObjectName("btn_buscar")
        self.btn_buscar.setMinimumWidth(80)
        self.btn_buscar.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        search_layout.addWidget(self.btn_buscar)

        self.btn_limpiar_busqueda = QPushButton("Limpiar")
        self.btn_limpiar_busqueda.setObjectName("btn_limpiar_busqueda")
        self.btn_limpiar_busqueda.setMinimumWidth(70)
        self.btn_limpiar_busqueda.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        search_layout.addWidget(self.btn_limpiar_busqueda)

        layout.addLayout(search_layout)
        
        # Tabla
        self.tabla = self._crear_tabla()
        layout.addWidget(self.tabla)
        
        # ── Barra de progreso estilo MacProgressBar ──────────
        self.barra_progreso = MacProgressBar()
        self.barra_progreso.setObjectName("historial_progress")
        self.barra_progreso.setRange(0, 100)
        self.barra_progreso.setValue(0)
        layout.addWidget(self.barra_progreso)

        # ── Fila de estado responsive ─────────────────────────
        status_frame = QFrame()
        status_frame.setObjectName("historial_status_bar")
        status_row = QHBoxLayout(status_frame)
        status_row.setContentsMargins(4, 4, 4, 0)
        status_row.setSpacing(8)

        self.status_bar = QLabel("✅ Listo")
        self.status_bar.setObjectName("status_bar")
        self.status_bar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        status_row.addWidget(self.status_bar)

        self.info_registros = QLabel("📦 0 guías")
        self.info_registros.setObjectName("info_registros")
        self.info_registros.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_registros.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        status_row.addWidget(self.info_registros)

        self.contador_tiempo = QLabel("⏱  00:00:00")
        self.contador_tiempo.setObjectName("contador_tiempo")
        self.contador_tiempo.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.contador_tiempo.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        status_row.addWidget(self.contador_tiempo)

        layout.addWidget(status_frame)
        
        return panel

    def _crear_toolbar(self):
        """Crea la toolbar superior en 2 filas para ser responsive."""
        from PySide6.QtWidgets import QSizePolicy as SP

        toolbar = QVBoxLayout()
        toolbar.setSpacing(4)

        # ── Fila 1: título + badge + tiempo real + actualizar ─────────
        fila1 = QHBoxLayout()
        fila1.setSpacing(6)

        self.titulo_label = QLabel("📋 REGISTRO DETALLADO")
        self.titulo_label.setObjectName("titulo_tabla")
        self.titulo_label.setSizePolicy(SP.Policy.Preferred, SP.Policy.Fixed)
        fila1.addWidget(self.titulo_label)

        self.lbl_tipo_activo = QLabel("")
        self.lbl_tipo_activo.setObjectName("lbl_tipo_activo")
        self.lbl_tipo_activo.setVisible(False)
        fila1.addWidget(self.lbl_tipo_activo)

        fila1.addStretch()

        self.cb_auto_actualizar = QCheckBox("🔄 Tiempo Real")
        self.cb_auto_actualizar.setObjectName("cb_auto_actualizar")
        self.cb_auto_actualizar.setChecked(True)
        self.cb_auto_actualizar.setSizePolicy(SP.Policy.Preferred, SP.Policy.Fixed)
        fila1.addWidget(self.cb_auto_actualizar)

        self.btn_actualizar = QPushButton("🔄 ACTUALIZAR")
        self.btn_actualizar.setObjectName("btn_actualizar")
        self.btn_actualizar.setToolTip("Actualizar datos ahora")
        self.btn_actualizar.setMinimumWidth(110)
        self.btn_actualizar.setSizePolicy(SP.Policy.Preferred, SP.Policy.Fixed)
        fila1.addWidget(self.btn_actualizar)

        toolbar.addLayout(fila1)

        # ── Fila 2: exportar + cerrar ─────────────────────────────────
        fila2 = QHBoxLayout()
        fila2.setSpacing(6)
        fila2.addStretch()

        self.btn_exportar_csv = QPushButton("📥 CSV")
        self.btn_exportar_csv.setObjectName("btn_exportar_csv")
        self.btn_exportar_csv.setToolTip("Exportar a CSV")
        self.btn_exportar_csv.setMinimumWidth(80)
        self.btn_exportar_csv.setSizePolicy(SP.Policy.Preferred, SP.Policy.Fixed)
        fila2.addWidget(self.btn_exportar_csv)

        self.btn_exportar_excel = QPushButton("📊 EXCEL")
        self.btn_exportar_excel.setObjectName("btn_exportar_excel")
        self.btn_exportar_excel.setToolTip("Exportar a Excel")
        self.btn_exportar_excel.setMinimumWidth(90)
        self.btn_exportar_excel.setSizePolicy(SP.Policy.Preferred, SP.Policy.Fixed)
        fila2.addWidget(self.btn_exportar_excel)

        self.btn_cerrar = QPushButton("✖ CERRAR")
        self.btn_cerrar.setObjectName("btn_cerrar_historial")
        self.btn_cerrar.setMinimumWidth(90)
        self.btn_cerrar.setSizePolicy(SP.Policy.Preferred, SP.Policy.Fixed)
        self.btn_cerrar.clicked.connect(self.accept)
        fila2.addWidget(self.btn_cerrar)

        toolbar.addLayout(fila2)

        return toolbar

    def _crear_tabla(self):
        """Crea la tabla de historial"""
        tabla = QTableWidget()
        tabla.setColumnCount(5)
        tabla.setHorizontalHeaderLabels(["📦 Guía", "📌 Estado", "📝 Resultado", "🌐 Navegador", "⏰ Fecha/Hora"])
        
        # Configurar tabla
        header = tabla.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        
        tabla.setAlternatingRowColors(True)
        tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tabla.setSortingEnabled(True)
        tabla.itemDoubleClicked.connect(self.copiar_guia)
        tabla.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        tabla.customContextMenuRequested.connect(self._on_context_menu)
        
        # El estilo de la tabla lo maneja _setup_styles globalmente
        tabla.setStyleSheet("")
        
        return tabla

    def _setup_styles(self):
        """Configura los estilos de la ventana"""
        c = theme.colors()
        self.setStyleSheet(theme.base_stylesheet() + f"""
            QDialog {{ background-color: {c['bg']}; }}

            /* ── Panels — glass cards (dark-mode aware) ── */
            QWidget#panel_izquierdo {{
                background-color: {c['surface']};
                border-radius: 14px;
                border: 1.5px solid {c['border']};
                border-bottom: 2px solid {c['border_strong']};
            }}
            QWidget#panel_derecho {{
                background-color: {c['surface']};
                border-radius: 14px;
                border: 1.5px solid {c['border']};
                border-bottom: 2px solid {c['border_strong']};
            }}

            /* ── Titles ── */
            QLabel#titulo_panel {{
                color: {c['text']};
                font-size: 11pt;
                font-weight: 700;
                padding: 8px 12px;
                letter-spacing: -0.2px;
            }}
            QLabel#titulo_tabla {{
                color: {c['text']};
                font-size: 11pt;
                font-weight: 700;
                letter-spacing: -0.2px;
            }}
            QLabel#lbl_tipo_activo {{
                color: white;
                background-color: {c['purple']};
                font-weight: 700;
                font-size: 9.5pt;
                padding: 3px 12px;
                border-radius: 12px;
                margin-left: 8px;
            }}

            /* ── Separators ── */
            QFrame#separador {{
                background-color: {c['sep']};
                max-height: 1px;
                border: none;
            }}

            /* ── Filter group ── */
            QGroupBox#filtros_group {{
                color: {c['group_title']};
                font-weight: 700;
                border: 1.5px solid {c['group_border']};
                border-radius: 12px;
                margin-top: 18px;
                padding-top: 14px;
                background-color: {c['group_bg']};
                font-size: 9pt;
            }}
            QGroupBox#filtros_group::title {{
                subcontrol-origin: margin;
                left: 14px;
                padding: 2px 8px;
                color: {c['group_title']};
                background-color: {c['group_bg']};
                border-radius: 4px;
            }}

            /* ── Status bar ── */
            QLabel#status_bar {{
                color: {c['success']};
                padding: 5px;
                font-weight: 600;
            }}
            QLabel#info_registros {{
                color: {c['text2']};
                padding: 5px 12px;
                background-color: {c['status_bg']};
                border-radius: 8px;
                font-weight: 600;
                font-size: 9.5pt;
                border: 1px solid {c['border']};
            }}
            QLabel#ultima_actualizacion {{
                color: {c['text3']};
                font-size: 8pt;
            }}

            /* ── Status bar inferior ── */
            QFrame#historial_status_bar {{
                background-color: transparent;
            }}
            QLabel#contador_tiempo {{
                color: {c['accent']};
                font-weight: 700;
                font-size: 10pt;
                padding: 4px 10px;
                background-color: {c['status_bg']};
                border-radius: 8px;
                border: 1px solid {c['border']};
            }}

            /* ── Checkboxes ── */
            QCheckBox#cb_seleccionar_todos {{
                color: {c['accent']};
                font-weight: 700;
                font-size: 10pt;
                spacing: 8px;
                padding: 6px 10px;
                background-color: {c['accent_light']};
                border-radius: 8px;
                border: 1px solid {c['accent']};
            }}
            QCheckBox#cb_auto_actualizar {{
                color: {c['accent']};
                font-weight: 600;
                spacing: 6px;
            }}

            /* ── Table — inherits from base but override here ── */
            QTableWidget {{
                border: 1.5px solid {c['border']};
                border-radius: 12px;
                font-size: 9.5pt;
            }}
            QTableWidget::item {{
                padding: 8px 12px;
                color: {c['text']};
                border-bottom: 1px solid {c['table_grid']};
            }}
            QTableWidget::item:selected {{
                background-color: {c['accent']};
                color: white;
            }}
            QHeaderView::section {{
                font-size: 9pt;
                font-weight: 700;
                letter-spacing: 0.3px;
                padding: 8px 12px;
            }}
            QTableCornerButton::section {{
                background-color: {c['table_hdr_bg']};
                border: none;
            }}

            /* ── Buttons ── */
            QPushButton#btn_aplicar_filtros {{
                background-color: {c['accent']};
                color: white;
                border: none;
                font-weight: 700;
            }}
            QPushButton#btn_aplicar_filtros:hover {{
                background-color: {c['accent_hover']};
            }}
            QPushButton#btn_aplicar_filtros:pressed {{
                background-color: {c['accent_press']};
            }}
            QPushButton#btn_limpiar_filtros {{
                color: {c['text3']};
                border-color: {c['border']};
            }}
            QPushButton#btn_actualizar {{
                background-color: {c['accent']};
                color: white;
                border: none;
                font-weight: 700;
            }}
            QPushButton#btn_actualizar:hover {{
                background-color: {c['accent_hover']};
            }}
            QPushButton#btn_exportar_csv {{
                color: {c['success']};
                border-color: {c['success']};
            }}
            QPushButton#btn_exportar_csv:hover {{
                background-color: {c['success_bg']};
                color: {c['success']};
            }}
            QPushButton#btn_exportar_excel {{
                color: {c['accent']};
                border-color: {c['accent']};
            }}
            QPushButton#btn_exportar_excel:hover {{
                background-color: {c['accent_light']};
                color: {c['accent']};
            }}
            QPushButton#btn_cerrar_historial {{
                background-color: {c['error']};
                color: white;
                border: none;
                font-weight: 700;
            }}
            QPushButton#btn_cerrar_historial:hover {{
                background-color: {c['error']};
                border: 2px solid {c['error']};
            }}
            QPushButton#btn_buscar {{
                background-color: {c['accent']};
                color: white;
                border: none;
                font-weight: 700;
            }}
            QPushButton#btn_buscar:hover {{ background-color: {c['accent_hover']}; }}
            QPushButton#btn_limpiar_busqueda {{
                color: {c['text3']};
                border-color: {c['border']};
            }}
        """)

    def _conectar_senales(self):
        """Conecta todas las señales"""
        # Checkbox "Seleccionar todos"
        self.cb_seleccionar_todos.stateChanged.connect(self.toggle_seleccionar_todos)
        
        # Checkboxes individuales
        for simbolo, cb in self.checkboxes.items():
            cb.stateChanged.connect(self.actualizar_estado_seleccion_todos)
            cb.stateChanged.connect(self.aplicar_filtros_multiple)
        
        # Otros filtros
        self.navegador_combo.currentTextChanged.connect(self.aplicar_filtros_multiple)
        self.buscar_input.currentTextChanged.connect(self.buscar_por_texto)
        self.buscar_input.editTextChanged.connect(self.buscar_por_texto)
        
        # Botones
        self.btn_buscar.clicked.connect(self.ejecutar_busqueda)
        self.btn_limpiar_busqueda.clicked.connect(self.limpiar_busqueda)
        self.btn_aplicar.clicked.connect(self.aplicar_filtros_multiple)
        self.btn_limpiar.clicked.connect(self.limpiar_todos_filtros)
        self.btn_exportar_csv.clicked.connect(self.exportar_csv)
        self.btn_exportar_excel.clicked.connect(self.exportar_excel)
        self.btn_actualizar.clicked.connect(self.solicitar_actualizacion_manual)
        
        # Checkbox auto-actualizar
        self.cb_auto_actualizar.stateChanged.connect(self.toggle_auto_actualizar)
        
        # Iniciar actualización automática por defecto
        self.toggle_auto_actualizar(Qt.CheckState.Checked.value)

    def toggle_seleccionar_todos(self, estado):
        """Selecciona o deselecciona todos los checkboxes de filtro"""
        seleccionar = estado == Qt.CheckState.Checked.value
        for cb in self.checkboxes.values():
            # Bloquear señales temporalmente para evitar recursión
            cb.blockSignals(True)
            cb.setChecked(seleccionar)
            cb.blockSignals(False)
        
        # Aplicar filtros después de cambiar todos
        self.aplicar_filtros_multiple()

    def actualizar_estado_seleccion_todos(self):
        """Actualiza el estado del checkbox 'Seleccionar todos'"""
        # Bloquear señales del checkbox "todos" para evitar recursión
        self.cb_seleccionar_todos.blockSignals(True)
        
        todos_checked = all(cb.isChecked() for cb in self.checkboxes.values())
        algun_checked = any(cb.isChecked() for cb in self.checkboxes.values())
        
        if todos_checked:
            self.cb_seleccionar_todos.setCheckState(Qt.CheckState.Checked)
        elif algun_checked:
            self.cb_seleccionar_todos.setCheckState(Qt.CheckState.PartiallyChecked)
        else:
            self.cb_seleccionar_todos.setCheckState(Qt.CheckState.Unchecked)
        
        self.cb_seleccionar_todos.blockSignals(False)

    def toggle_auto_actualizar(self, estado):
        """Activa o desactiva la actualización automática en tiempo real"""
        self.actualizacion_automatica = estado == Qt.CheckState.Checked.value
        
        if self.actualizacion_automatica:
            self.timer_actualizacion.start()
            self.status_bar.setText("🔄 Modo tiempo real activado (c/2s)")
            # Forzar una actualización inmediata al activar
            self.solicitar_actualizacion_manual()
        else:
            self.timer_actualizacion.stop()
            self.status_bar.setText("✅ Modo tiempo real desactivado")

    def _actualizacion_automatica_timeout(self):
        """Timeout del timer de actualización automática"""
        self.solicitar_actualizacion_manual()

    def actualizar_contador_tiempo(self):
        """Actualiza el contador con tiempo restante estimado (basado en velocidad de proceso)"""
        procesadas = len(self.datos_completos)
        total = self.total_esperado

        if total > 0 and procesadas >= total:
            self.contador_tiempo.setText("✅ Completado")
            return

        if total > 0 and procesadas > 0:
            elapsed_s = (datetime.now() - self.tiempo_inicio).total_seconds()
            if elapsed_s > 0:
                rate = procesadas / elapsed_s          # guías/seg
                pendientes = total - procesadas
                restantes_s = pendientes / rate
                horas = int(restantes_s // 3600)
                minutos = int((restantes_s % 3600) // 60)
                segundos = int(restantes_s % 60)
                self.contador_tiempo.setText(f"⏳ {horas:02d}:{minutos:02d}:{segundos:02d}")
                return

        if total > 0 and procesadas == 0:
            self.contador_tiempo.setText("⏳ Calculando...")
            return

        # Sin total conocido: mostrar tiempo transcurrido
        elapsed = datetime.now() - self.tiempo_inicio
        horas = int(elapsed.total_seconds() // 3600)
        minutos = int((elapsed.total_seconds() % 3600) // 60)
        segundos = int(elapsed.total_seconds() % 60)
        self.contador_tiempo.setText(f"⏱  {horas:02d}:{minutos:02d}:{segundos:02d}")

    def solicitar_actualizacion_manual(self):
        """Solicita actualización manual de datos al padre"""
        self.status_bar.setText("🔄 Actualizando datos en tiempo real...")
        
        if self.parent() and hasattr(self.parent(), 'obtener_datos_historial'):
            try:
                nuevos_datos = self.parent().obtener_datos_historial()
                if nuevos_datos is not None:
                    # Verificar si hay cambios
                    if len(nuevos_datos) != len(self.datos_completos) or nuevos_datos != self.datos_completos:
                        self.actualizar_historial(nuevos_datos)
                        hora_actual = datetime.now().strftime('%H:%M:%S')
                        self.status_bar.setText(f"✅ Actualizado: {hora_actual}")
                        # Animar los círculos cuando hay cambios
                        self._animar_estadisticas()
                    else:
                        self.status_bar.setText("✅ Datos al día")
                else:
                    self.status_bar.setText("⚠️ No se pudieron obtener datos")
            except Exception as e:
                self.status_bar.setText(f"❌ Error al actualizar: {str(e)}")
        else:
            self.status_bar.setText("⚠️ Método obtener_datos_historial no disponible")

    def _animar_estadisticas(self):
        """Efecto visual para indicar que las estadísticas se actualizaron"""
        # Cambiar temporalmente el estilo de los círculos
        for stat in self.stats.values():
            stat.setStyleSheet("border: 2px solid white;")
            QTimer.singleShot(200, lambda s=stat: s.setStyleSheet(""))

    def actualizar_historial(self, datos, tipo_activo=None):
        """Actualiza el historial con nuevos datos y las estadísticas en tiempo real"""
        self.datos_completos = datos.copy()
        self.datos_filtrados = datos.copy()
        if tipo_activo:
            self.lbl_tipo_activo.setText(f"  📌 Desviación activa: Tipo {tipo_activo}  ")
            self.lbl_tipo_activo.setVisible(True)
        self._actualizar_estadisticas()
        self._actualizar_sugerencias_busqueda()
        self.aplicar_filtros_multiple()

    def set_total_esperado(self, total: int):
        """Establece el total de guías esperadas para la barra de progreso."""
        self.total_esperado = max(total, 0)
        self._actualizar_barra_progreso()

    def _actualizar_barra_progreso(self):
        """Actualiza la barra de progreso y el resumen de guías."""
        procesadas = len(self.datos_completos)
        total = self.total_esperado if self.total_esperado > 0 else procesadas
        pct = int(procesadas / total * 100) if total > 0 else 0
        self.barra_progreso.setValue(pct)
        pendientes = max(total - procesadas, 0)
        if self.total_esperado > 0:
            self.info_registros.setText(
                f"📦 {procesadas}/{total}  ·  {pendientes} pendientes"
            )
        else:
            self.info_registros.setText(f"📦 {procesadas} guías")

    def _actualizar_estadisticas(self):
        """Actualiza las estadísticas circulares en tiempo real"""
        total = len(self.datos_completos)
        self.stats["total"].actualizar_valor(total)

        conteos = {"exitosas": 0, "ent": 0, "errores": 0, "advertencias": 0}

        for dato in self.datos_completos:
            estado = dato[1]
            if "✅" in estado:
                conteos["exitosas"] += 1
            if "📦" in estado:
                conteos["ent"] += 1
            if "❌" in estado:
                conteos["errores"] += 1
            if "⚠️" in estado:
                conteos["advertencias"] += 1

        for key, valor in conteos.items():
            self.stats[key].actualizar_valor(valor)

        self._actualizar_barra_progreso()

    def _actualizar_sugerencias_busqueda(self):
        """Actualiza las sugerencias del combo de búsqueda y el filtro de navegadores"""
        self.buscar_input.blockSignals(True)
        self.buscar_input.clear()
        guias = sorted(set(d[0] for d in self.datos_completos))
        self.buscar_input.addItems(guias)
        self.buscar_input.setCurrentText(self.texto_busqueda)
        self.buscar_input.blockSignals(False)

        # Actualizar opciones de navegador con los valores reales del historial
        navegador_actual = self.navegador_combo.currentText()
        self.navegador_combo.blockSignals(True)
        self.navegador_combo.clear()
        navegadores = sorted(set(d[3] for d in self.datos_completos))
        self.navegador_combo.addItem("Todos")
        self.navegador_combo.addItems(navegadores)
        # Restaurar selección previa si aún existe
        idx = self.navegador_combo.findText(navegador_actual)
        self.navegador_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.navegador_combo.blockSignals(False)

    def aplicar_filtros_multiple(self):
        """Aplica múltiples filtros simultáneamente"""
        # Recoger estado de checkboxes
        estados_seleccionados = []
        for simbolo, cb in self.checkboxes.items():
            if cb.isChecked():
                estados_seleccionados.append(simbolo)
        
        # Si no hay checkboxes seleccionados, mostrar todos
        if not estados_seleccionados:
            self.datos_filtrados = self.datos_completos.copy()
        else:
            # Filtrar por estados seleccionados
            self.datos_filtrados = [
                d for d in self.datos_completos 
                if any(estado in d[1] for estado in estados_seleccionados)
            ]
        
        # Aplicar filtro de navegador
        navegador = self.navegador_combo.currentText()
        if navegador != "Todos":
            self.datos_filtrados = [
                d for d in self.datos_filtrados 
                if navegador in d[3]
            ]
        
        # Aplicar búsqueda por texto
        if self.texto_busqueda:
            self.datos_filtrados = [
                d for d in self.datos_filtrados
                if self.texto_busqueda.lower() in d[0].lower()
            ]
        
        self._actualizar_vista()
        self.filtros_actualizados.emit()

    def buscar_por_texto(self, texto):
        """Actualiza el texto de búsqueda"""
        self.texto_busqueda = texto

    def ejecutar_busqueda(self):
        """Ejecuta la búsqueda con el texto actual"""
        self.aplicar_filtros_multiple()

    def limpiar_busqueda(self):
        """Limpia el campo de búsqueda"""
        self.buscar_input.setCurrentText("")
        self.texto_busqueda = ""
        self.aplicar_filtros_multiple()

    def limpiar_todos_filtros(self):
        """Limpia todos los filtros aplicados"""
        # Desmarcar todos los checkboxes
        self.cb_seleccionar_todos.setChecked(False)
        for cb in self.checkboxes.values():
            cb.setChecked(False)
        
        # Resetear combo
        self.navegador_combo.setCurrentText("Todos")
        
        # Limpiar búsqueda
        self.buscar_input.setCurrentText("")
        self.texto_busqueda = ""
        
        # Restaurar datos
        self.datos_filtrados = self.datos_completos.copy()
        self._actualizar_vista()
        
        self.status_bar.setText("✅ Filtros limpiados")

    def _actualizar_vista(self):
        """Actualiza la vista de la tabla"""
        datos_ordenados = sorted(self.datos_filtrados, key=lambda x: x[4], reverse=True)
        self.tabla.setRowCount(len(datos_ordenados))
        self.tabla.setSortingEnabled(False)
        
        for i, (guia, estado, resultado, nav, fecha) in enumerate(datos_ordenados):
            self._agregar_fila(i, guia, estado, resultado, nav, fecha)
        
        self._ajustar_columnas()
        self.tabla.setSortingEnabled(True)
        
        # Actualizar info de registros
        total = len(self.datos_completos)
        filtrados = len(self.datos_filtrados)
        if total == filtrados:
            self.info_registros.setText(f"{total} registros")
        else:
            self.info_registros.setText(f"{filtrados} de {total} registros")

    def _agregar_fila(self, fila, guia, estado, resultado, nav, fecha):
        """Agrega una fila a la tabla con formato"""
        BLANCO = QColor("#ecf0f1")
        font_bold = QFont()
        font_bold.setBold(True)

        # Guía
        item_guia = QTableWidgetItem(guia)
        item_guia.setForeground(BLANCO)
        item_guia.setFont(font_bold)
        item_guia.setToolTip(f"Haz doble clic para copiar: {guia}")
        self.tabla.setItem(fila, 0, item_guia)

        # Estado
        item_estado = QTableWidgetItem(estado)
        config_estado = self._get_estado_config(estado)
        item_estado.setForeground(QColor(config_estado['color']))
        item_estado.setBackground(QColor(config_estado['background']))
        item_estado.setToolTip(config_estado['tooltip'])
        item_estado.setFont(font_bold)
        self.tabla.setItem(fila, 1, item_estado)

        # Resultado
        item_resultado = QTableWidgetItem(resultado)
        config_resultado = self._get_resultado_config(resultado)
        item_resultado.setForeground(QColor(config_resultado['color']))
        item_resultado.setBackground(QColor(config_resultado['background']))
        item_resultado.setToolTip(config_resultado['tooltip'])
        item_resultado.setFont(font_bold)
        self.tabla.setItem(fila, 2, item_resultado)

        # Navegador
        item_nav = QTableWidgetItem(nav)
        item_nav.setForeground(BLANCO)
        item_nav.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tabla.setItem(fila, 3, item_nav)

        # Fecha
        item_fecha = QTableWidgetItem(fecha)
        item_fecha.setForeground(BLANCO)
        item_fecha.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tabla.setItem(fila, 4, item_fecha)

    def _get_estado_config(self, estado):
        """Retorna configuración de color para el estado"""
        configs = {
            "✅": {'color': "#27ae60", 'background': "#1a472a", 'tooltip': "✅ Procesada exitosamente"},
            "📦": {'color': "#f39c12", 'background': "#4a3b1a", 'tooltip': "📦 Guía entregada (ENT)"},
            "❌": {'color': "#e74c3c", 'background': "#4a1a1a", 'tooltip': "❌ Error en procesamiento"},
            "⚠️": {'color': "#e67e22", 'background': "#4a2a1a", 'tooltip': "⚠️ Advertencia - Verificar"}
        }
        for key, config in configs.items():
            if key in estado:
                return config
        return {'color': "#95a5a6", 'background': "#2c3e50", 'tooltip': "Estado desconocido"}

    def _get_resultado_config(self, resultado):
        """Retorna configuración de color para el resultado"""
        if "ENT" in resultado:
            return {'color': "#f39c12", 'background': "#4a3b1a", 'tooltip': "📦 Guía con estado ENT"}
        elif "ADVERTENCIA" in resultado or "NO CONFIRMADO" in resultado:
            return {'color': "#e67e22", 'background': "#4a2a1a", 'tooltip': "⚠️ Completado con advertencias"}
        elif "ERROR" in resultado:
            return {'color': "#e74c3c", 'background': "#4a1a1a", 'tooltip': "❌ Error en procesamiento"}
        elif "COMPLETADO" in resultado or resultado.startswith("Tipo ") or resultado.startswith("Incidencia "):
            return {'color': "#27ae60", 'background': "#1a472a", 'tooltip': "✅ Procesado correctamente"}
        elif "SIN RESULTADOS" in resultado:
            return {'color': "#e74c3c", 'background': "#4a1a1a", 'tooltip': "❌ Guía no encontrada"}
        return {'color': "#95a5a6", 'background': "#2c3e50", 'tooltip': resultado}

    def _ajustar_columnas(self):
        """Ajusta el ancho de las columnas"""
        self.tabla.resizeColumnsToContents()
        if self.tabla.columnWidth(0) < 100:
            self.tabla.setColumnWidth(0, 100)
        if self.tabla.columnWidth(1) < 100:
            self.tabla.setColumnWidth(1, 100)
        if self.tabla.columnWidth(3) < 80:
            self.tabla.setColumnWidth(3, 80)
        if self.tabla.columnWidth(4) < 130:
            self.tabla.setColumnWidth(4, 130)

    def copiar_guia(self, item):
        """Copia la guía al portapapeles al hacer doble clic"""
        if item.column() == 0:
            guia = item.text()
            QApplication.clipboard().setText(guia)
            item.setSelected(True)
            self.status_bar.setText(f"✅ Guía '{guia}' copiada al portapapeles")
            QTimer.singleShot(3000, lambda: self.status_bar.setText("✅ Listo"))

    def _on_context_menu(self, pos):
        """Muestra menú contextual con opciones según el estado de la fila."""
        idx = self.tabla.indexAt(pos)
        if not idx.isValid():
            return

        fila = idx.row()
        item_guia   = self.tabla.item(fila, 0)
        item_estado = self.tabla.item(fila, 1)
        if not item_guia or not item_estado:
            return

        guia   = item_guia.text()
        estado = item_estado.text()
        c = theme.colors()

        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background-color: {c['surface']};
                border: 1.5px solid {c['border']};
                border-radius: 10px;
                padding: 4px;
                color: {c['text']};
                font-size: 9.5pt;
            }}
            QMenu::item {{
                padding: 7px 20px 7px 14px;
                border-radius: 6px;
            }}
            QMenu::item:selected {{
                background-color: {c['accent_light']};
                color: {c['accent']};
            }}
            QMenu::item:disabled {{
                color: {c['text3']};
            }}
            QMenu::separator {{
                height: 1px;
                background: {c['border']};
                margin: 4px 8px;
            }}
        """)

        # Copiar guía — siempre disponible
        accion_copiar = menu.addAction(f"📋  Copiar guía  {guia}")

        menu.addSeparator()

        # Reprocesar — solo para errores
        es_error = "❌" in estado
        accion_reprocesar = menu.addAction("🔄  Reprocesar esta guía")
        accion_reprocesar.setEnabled(es_error)

        accion = menu.exec(self.tabla.viewport().mapToGlobal(pos))

        if accion == accion_copiar:
            QApplication.clipboard().setText(guia)
            self.status_bar.setText(f"✅ Guía '{guia}' copiada al portapapeles")
            QTimer.singleShot(3000, lambda: self.status_bar.setText("✅ Listo"))
        elif accion == accion_reprocesar and es_error:
            self.reprocesar_guia.emit(guia)

    def exportar_csv(self):
        """Exporta los datos filtrados a CSV"""
        try:
            if not self.datos_filtrados:
                QMessageBox.warning(self, "⚠️ Advertencia", "No hay datos para exportar")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_nombre = f"historial_alertran_{timestamp}"
            
            ruta_completa = self.file_utils.generar_nombre_unico(
                self.carpeta_descargas, base_nombre, "csv"
            )
            
            with open(ruta_completa, 'w', encoding='utf-8-sig') as f:
                f.write("Guía,Estado,Resultado,Navegador,Fecha\n")
                for i in range(self.tabla.rowCount()):
                    fila = [
                        self.tabla.item(i, 0).text(),
                        self.tabla.item(i, 1).text(),
                        self.tabla.item(i, 2).text(),
                        self.tabla.item(i, 3).text(),
                        self.tabla.item(i, 4).text()
                    ]
                    f.write(','.join(fila) + '\n')
            
            QMessageBox.information(
                self, "✅ Exportación Exitosa", 
                f"Archivo guardado en:\n{ruta_completa}"
            )
            self.status_bar.setText(f"✅ Exportado: {Path(ruta_completa).name}")
            
        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"No se pudo exportar:\n{str(e)}")

    def exportar_excel(self):
        """Exporta los datos filtrados a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            if not self.datos_filtrados:
                QMessageBox.warning(self, "⚠️ Advertencia", "No hay datos para exportar")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_nombre = f"historial_alertran_{timestamp}"
            
            ruta_completa = self.file_utils.generar_nombre_unico(
                self.carpeta_descargas, base_nombre, "xlsx"
            )
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial Alertran"
            
            # Encabezados
            headers = ["Guía", "Estado", "Resultado", "Navegador", "Fecha"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for row in range(self.tabla.rowCount()):
                for col in range(self.tabla.columnCount()):
                    item = self.tabla.item(row, col)
                    if item:
                        cell = ws.cell(row=row + 2, column=col + 1, value=item.text())
                        
                        # Aplicar colores según estado en columna 2
                        if col == 1:
                            if "✅" in item.text():
                                cell.fill = PatternFill(start_color="1A472A", end_color="1A472A", fill_type="solid")
                                cell.font = Font(color="27AE60")
                            elif "📦" in item.text():
                                cell.fill = PatternFill(start_color="4A3B1A", end_color="4A3B1A", fill_type="solid")
                                cell.font = Font(color="F39C12")
                            elif "❌" in item.text():
                                cell.fill = PatternFill(start_color="4A1A1A", end_color="4A1A1A", fill_type="solid")
                                cell.font = Font(color="E74C3C")
                            elif "⚠️" in item.text():
                                cell.fill = PatternFill(start_color="4A2A1A", end_color="4A2A1A", fill_type="solid")
                                cell.font = Font(color="E67E22")
            
            # Ajustar ancho de columnas
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 20
            
            wb.save(ruta_completa)
            
            QMessageBox.information(
                self, "✅ Exportación Exitosa", 
                f"Archivo guardado en:\n{ruta_completa}"
            )
            self.status_bar.setText(f"✅ Exportado: {Path(ruta_completa).name}")
            
        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"No se pudo exportar a Excel:\n{str(e)}")

    def closeEvent(self, event):
        sm = SettingsManager.get_instance()
        sm.set("historial_window_geometry", self.saveGeometry().toBase64().data().decode())
        sm.save()
        super().closeEvent(event)