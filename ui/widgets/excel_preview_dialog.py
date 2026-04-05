# ui/widgets/excel_preview_dialog.py
"""
Diálogo de vista previa del archivo Excel antes de iniciar el proceso.
Muestra las primeras filas y el total de guías para que el usuario confirme.
"""
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QColor
from pathlib import Path
from openpyxl import load_workbook
from utils import theme


class ExcelPreviewDialog(QDialog):
    """Vista previa del Excel: muestra N filas y el total de guías."""

    MAX_FILAS = 25

    def __init__(self, ruta: str, total_guias: int, parent=None):
        super().__init__(parent)
        self.ruta        = Path(ruta)
        self.total_guias = total_guias
        self.setWindowTitle("📊 Vista previa — Archivo Excel")
        self.setMinimumSize(460, 480)
        self.setSizeGripEnabled(True)
        self.setModal(True)
        self._setup_ui()
        self._setup_styles()
        self._cargar_datos()
        from utils import theme
        theme.signals.changed.connect(self._setup_styles)

    # ── UI ───────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 0)

        # Header
        header = QFrame()
        header.setObjectName("prev_header")
        header.setFixedHeight(58)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(20, 0, 20, 0)

        titulo = QLabel(f"📄  {self.ruta.name}")
        titulo.setObjectName("prev_titulo")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(titulo)
        layout.addWidget(header)

        # Body
        body = QFrame()
        body.setObjectName("prev_body")
        body_layout = QVBoxLayout(body)
        body_layout.setSpacing(10)
        body_layout.setContentsMargins(20, 16, 20, 16)

        # Badge de total
        badge = QLabel(f"✅  {self.total_guias:,} guías encontradas")
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        badge.setObjectName("prev_badge")
        body_layout.addWidget(badge)

        sub = QLabel(f"Mostrando las primeras {self.MAX_FILAS} filas")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setObjectName("prev_sub")
        body_layout.addWidget(sub)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setObjectName("prev_sep")
        body_layout.addWidget(sep)

        # Tabla
        self.tabla = QTableWidget()
        self.tabla.setObjectName("prev_tabla")
        self.tabla.setColumnCount(1)
        self.tabla.setHorizontalHeaderLabels(["Número de Guía"])
        self.tabla.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setAlternatingRowColors(True)
        body_layout.addWidget(self.tabla)

        layout.addWidget(body, stretch=1)

        # Botones
        btn_bar = QFrame()
        btn_bar.setObjectName("prev_btn_bar")
        btn_row = QHBoxLayout(btn_bar)
        btn_row.setContentsMargins(20, 12, 20, 12)
        btn_row.setSpacing(10)

        self.btn_cancelar = QPushButton("❌  Cancelar")
        self.btn_cancelar.setObjectName("btn_prev_cancel")
        self.btn_cancelar.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_cancelar)

        btn_row.addStretch()

        self.btn_continuar = QPushButton("✅  Continuar con este archivo")
        self.btn_continuar.setObjectName("btn_prev_ok")
        self.btn_continuar.clicked.connect(self.accept)
        btn_row.addWidget(self.btn_continuar)

        layout.addWidget(btn_bar)

    def _setup_styles(self):
        c = theme.colors()
        self.setStyleSheet(theme.base_stylesheet() + f"""
            QDialog {{ background-color: {c['bg']}; }}

            /* ── Header ── */
            QFrame#prev_header {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {c['accent']}, stop:1 {c['purple']});
                border-bottom: 2px solid {c['accent']};
            }}
            QLabel#prev_titulo {{
                color: white;
                font-size: 11pt;
                font-weight: 700;
            }}

            /* ── Body ── */
            QFrame#prev_body {{ background-color: {c['bg']}; }}

            QLabel#prev_badge {{
                color: {c['success']};
                font-size: 11pt;
                font-weight: bold;
                padding: 6px 14px;
                background-color: {c['success_bg']};
                border: 1px solid {c['success']};
                border-radius: 8px;
            }}
            QLabel#prev_sub {{
                color: {c['text3']};
                font-size: 9pt;
            }}
            QFrame#prev_sep {{
                background-color: {c['border']};
                max-height: 1px;
            }}

            /* ── Tabla ── */
            QTableWidget#prev_tabla {{
                background-color: {c['surface']};
                alternate-background-color: {c['surface2']};
                color: {c['text']};
                gridline-color: {c['border']};
                font-size: 10pt;
                border: 1.5px solid {c['border']};
                border-radius: 8px;
            }}
            QHeaderView::section {{
                background-color: {c['surface2']};
                color: {c['text2']};
                font-weight: bold;
                padding: 7px;
                border: none;
                border-bottom: 1px solid {c['border']};
            }}
            QTableWidget::item {{ padding: 5px; }}
            QTableWidget::item:selected {{
                background-color: {c['accent_light']};
                color: {c['accent']};
            }}

            /* ── Botones ── */
            QFrame#prev_btn_bar {{
                background-color: {c['action_bg']};
                border-top: 1px solid {c['border']};
            }}
            QPushButton {{
                padding: 9px 20px;
                border-radius: 10px;
                font-weight: bold;
                font-size: 10pt;
                min-height: 36px;
            }}
            QPushButton#btn_prev_ok {{
                background-color: {c['success']};
                color: white;
                border: none;
            }}
            QPushButton#btn_prev_ok:hover {{
                border: 2px solid {c['success']};
            }}
            QPushButton#btn_prev_ok:pressed {{
                background-color: {c['success_bg']};
                color: {c['success']};
                border: 2px solid {c['success']};
            }}
            QPushButton#btn_prev_cancel {{
                background-color: transparent;
                color: {c['text2']};
                border: 1.5px solid {c['border']};
            }}
            QPushButton#btn_prev_cancel:hover {{
                background-color: {c['surface2']};
                border-color: {c['border_strong']};
            }}
        """)

    # ── Datos ────────────────────────────────────────────────────────────────

    def _cargar_datos(self):
        c = theme.colors()
        try:
            wb = load_workbook(self.ruta, read_only=True, data_only=True)
            ws = wb.active
            filas = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    filas.append(str(row[0]).strip())
                    if len(filas) >= self.MAX_FILAS:
                        break
            wb.close()

            self.tabla.setRowCount(len(filas))
            for i, guia in enumerate(filas):
                item = QTableWidgetItem(guia)
                item.setForeground(QColor(c['text']))
                self.tabla.setItem(i, 0, item)

        except Exception as e:
            self.tabla.setRowCount(1)
            err = QTableWidgetItem(f"Error al leer: {str(e)}")
            err.setForeground(QColor(c['error']))
            self.tabla.setItem(0, 0, err)
